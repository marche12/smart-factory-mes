#!/usr/bin/env python3
"""
얼마에요2E에서 내보낸 엑셀/CSV를 팩플로우 형식으로 변환하여 DB에 넣는 스크립트.

사용법:
  1. migrate/exported/ 폴더에 엑셀(.xlsx) 또는 CSV 파일을 넣기
     - 얼마에요에서 거래처/품목 목록을 엑셀로 내보내기
  2. python3 02_import_to_packflow.py 실행

  수동 입력: python3 02_import_to_packflow.py --manual
"""
import csv
import json
import os
import sys
import uuid
import sqlite3
from datetime import datetime

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
EXPORT_DIR = os.path.join(SCRIPT_DIR, "exported")
DB_PATH = os.path.join(PROJECT_DIR, "data", "mes.db")


def gid():
    return str(uuid.uuid4())[:8]


def now_str():
    return datetime.now().isoformat()


def today():
    return datetime.now().strftime("%Y-%m-%d")


def read_csv(filename):
    """CSV 파일 읽기"""
    path = os.path.join(EXPORT_DIR, filename)
    if not os.path.exists(path):
        print(f"  [SKIP] {filename} 파일 없음")
        return []
    with open(path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        return list(reader)


def read_xlsx_auto():
    """exported/ 폴더의 모든 .xlsx 파일을 읽어 거래처/품목을 자동 판별"""
    try:
        import openpyxl
    except ImportError:
        print("openpyxl 설치 필요: pip3 install openpyxl")
        return [], []

    xlsx_files = [f for f in os.listdir(EXPORT_DIR) if f.endswith(('.xlsx', '.xls'))]
    if not xlsx_files:
        return [], []

    customers = []
    items = []

    for fname in xlsx_files:
        path = os.path.join(EXPORT_DIR, fname)
        print(f"  엑셀 읽기: {fname}")
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except Exception as e:
            print(f"    [ERROR] {e}")
            continue

        for sheet in wb.sheetnames:
            ws = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                continue

            # 헤더 행 찾기 (처음 나오는 텍스트 행)
            header_idx = 0
            for i, row in enumerate(rows):
                vals = [str(v or '').strip() for v in row]
                if any(v for v in vals):
                    header_idx = i
                    break

            headers = [str(v or '').strip().replace('\n', ' ') for v in rows[header_idx]]
            headers_lower = [h.lower() for h in headers]
            data_rows = rows[header_idx + 1:]

            # 거래처 판별: '거래처명', '상호', '사업자' 등의 컬럼이 있으면
            is_customer = any(k in ''.join(headers) for k in ['거래처', '상호', '사업자', '매출처', '매입처', '업체명'])
            # 품목 판별: '품목', '제품', '품명', '규격' 등
            is_item = any(k in ''.join(headers) for k in ['품목', '제품', '품명', '코드', '바코드'])

            if is_customer:
                print(f"    → 거래처 시트 감지: [{sheet}] ({len(data_rows)}행)")
                for row in data_rows:
                    vals = {headers[j]: (str(row[j]) if j < len(row) and row[j] is not None else '') for j in range(len(headers))}
                    # 거래처명 찾기
                    nm = ''
                    for key in ['거래처명', '상호', '거래처', '업체명', '회사명', '매출처', '매입처']:
                        if key in vals and vals[key].strip():
                            nm = vals[key].strip()
                            break
                    if not nm:
                        # 첫 번째 비어있지 않은 열
                        for v in row:
                            if v and str(v).strip() and not str(v).replace('-','').replace(' ','').isdigit():
                                nm = str(v).strip()
                                break
                    if not nm or nm == 'None':
                        continue
                    customers.append({
                        "nm": nm,
                        "bizNo": _find_val(vals, ['사업자번호', '사업자', '등록번호']),
                        "ceo": _find_val(vals, ['대표자', '대표', '대표자명']),
                        "bizType": _find_val(vals, ['업태']),
                        "bizClass": _find_val(vals, ['종목', '업종']),
                        "addr": _find_val(vals, ['주소', '소재지', '사업장주소']),
                        "tel": _find_val(vals, ['전화', '전화번호', '연락처', 'TEL']),
                        "fax": _find_val(vals, ['팩스', 'FAX']),
                        "email": _find_val(vals, ['이메일', 'E-mail', 'email']),
                        "customerType": _find_val(vals, ['유형', '구분', '분류']),
                        "note": _find_val(vals, ['비고', '메모', '참고']),
                    })

            elif is_item:
                print(f"    → 품목 시트 감지: [{sheet}] ({len(data_rows)}행)")
                for row in data_rows:
                    vals = {headers[j]: (str(row[j]) if j < len(row) and row[j] is not None else '') for j in range(len(headers))}
                    nm = ''
                    for key in ['품목명', '품명', '제품명', '제품', '품목', '상품명']:
                        if key in vals and vals[key].strip():
                            nm = vals[key].strip()
                            break
                    if not nm or nm == 'None':
                        continue
                    price_str = _find_val(vals, ['단가', '판매단가', '기본단가', '가격'])
                    try:
                        price = int(float(price_str.replace(',', ''))) if price_str else 0
                    except (ValueError, TypeError):
                        price = 0
                    items.append({
                        "nm": nm,
                        "code": _find_val(vals, ['코드', '품목코드', '제품코드', '품번']),
                        "spec": _find_val(vals, ['규격', '사양', '스펙']),
                        "price": price,
                        "note": _find_val(vals, ['비고', '메모', '참고']),
                    })
            else:
                print(f"    → [{sheet}] 거래처/품목 판별 불가 (컬럼: {', '.join(headers[:5])}...)")

    print(f"  총 거래처 {len(customers)}건, 품목 {len(items)}건 읽음")
    return customers, items


def _find_val(vals_dict, keys):
    """여러 가능한 키 중 값이 있는 것을 반환"""
    for k in keys:
        if k in vals_dict and vals_dict[k].strip() and vals_dict[k].strip() != 'None':
            return vals_dict[k].strip()
    return ''


def get_db_data(db, key):
    """팩플로우 DB에서 기존 데이터 읽기"""
    row = db.execute("SELECT value FROM data_store WHERE key=?", [key]).fetchone()
    if row:
        return json.loads(row[0])
    return []


def set_db_data(db, key, data):
    """팩플로우 DB에 데이터 쓰기"""
    db.execute(
        "INSERT OR REPLACE INTO data_store(key, value) VALUES(?, ?)",
        [key, json.dumps(data, ensure_ascii=False)]
    )


def infer_customer_type(*texts, receivable=0, payable=0):
    """얼마에요 거래처 유형/잔액 문구를 팩플로우 cType으로 변환."""
    s = "".join(str(t or "") for t in texts).lower().replace(" ", "")
    has_sales = any(k in s for k in ["매출", "판매", "수금", "받을", "customer", "sales"])
    has_purchase = any(k in s for k in ["매입", "구매", "외주", "협력", "공급", "지급", "줄돈", "vendor", "purchase"])
    if any(k in s for k in ["both", "양쪽", "겸", "공통", "매출매입", "매입매출"]) or (has_sales and has_purchase):
        return "both"
    if receivable and payable:
        return "both"
    if has_purchase or payable:
        return "purchase"
    if has_sales or receivable:
        return "sales"
    return "sales"


def merge_customer_type(current, evidence):
    current = current or "sales"
    if not evidence or current == evidence:
        return current
    if "both" in (current, evidence):
        return "both"
    if {current, evidence} == {"sales", "purchase"}:
        return "both"
    return evidence or current


def convert_customers(rows):
    """얼마에요 거래처 → 팩플로우 cli 형식 (Slip 스키마 기준 컬럼 호환)"""
    clients = []
    seen = set()
    for r in rows:
        nm = (r.get("nm") or "").strip()
        if not nm or nm in seen:
            continue
        seen.add(nm)

        ct = (r.get("customerType") or "").strip()
        note = (r.get("note") or "").strip()
        memo = (r.get("memo") or "").strip()
        full_note = (note + (("\n" + memo) if memo else "")).strip()
        ctype = infer_customer_type(ct, full_note)

        # 전화: CompanyTelephone > Handphone > HomeTelephone 순으로 우선 적용
        tel = (r.get("tel") or "").strip()
        handphone = (r.get("handphone") or "").strip()
        hometel = (r.get("hometel") or "").strip()
        primary_tel = tel or handphone or hometel

        clients.append({
            "id": gid(),
            "nm": nm,
            "bizNo": (r.get("bizNo") or "").strip(),
            "ceo": (r.get("ceo") or "").strip(),
            "bizType": (r.get("bizType") or "").strip(),
            "bizClass": (r.get("bizClass") or "").strip(),
            "addr": (r.get("addr") or "").strip(),
            "tel": primary_tel,
            "telCompany": tel,
            "telHandphone": handphone,
            "telHome": hometel,
            "fax": (r.get("fax") or "").strip(),
            "email": (r.get("email") or "").strip(),
            "zipcode": (r.get("zipcode") or "").strip(),
            "cType": ctype,
            "cTypeRaw": ct,
            "customerType": ct,
            "contactNm": (r.get("contactNm") or "").strip(),
            "contactTel": (r.get("contactTel") or "").strip(),
            "note": full_note,
            "almayoBookId": (r.get("bookId") or "").strip(),
            "almayoBookNo": (r.get("bookNo") or "").strip(),
            "ps": "",
            "cat": now_str(),
            "_src": "almayo"
        })
    return clients


def convert_items(rows, clients):
    """얼마에요 품목 → 팩플로우 prod 형식"""
    products = []
    seen = set()
    # 기본 공정 (패키지 제조)
    default_procs = [
        {"nm": "인쇄", "tp": "out"},
        {"nm": "코팅", "tp": "in"},
        {"nm": "톰슨", "tp": "in"},
        {"nm": "접착", "tp": "in"},
    ]

    for r in rows:
        nm = (r.get("nm") or "").strip()
        if not nm or nm in seen:
            continue
        seen.add(nm)

        price = 0
        try:
            price = int(float(r.get("price") or 0))
        except (ValueError, TypeError):
            pass

        products.append({
            "id": gid(),
            "nm": nm,
            "code": (r.get("code") or "").strip(),
            "cnm": "",  # 거래처 연결은 WO 등록 시 자동
            "cid": "",
            "spec": (r.get("spec") or "").strip(),
            "price": price,
            "paper": "",
            "papers": [],
            "fabric": "",
            "fabrics": [],
            "procs": default_procs,
            "ps": "",
            "gold": "",
            "mold": "",
            "hand": "",
            "nt": (r.get("note") or "").strip(),
            "caut": "",
            "cat": now_str(),
            "_src": "almayo"
        })
    return products


def convert_company(rows):
    """얼마에요 회사정보 → 팩플로우 co 형식"""
    if not rows:
        return None
    r = rows[0]
    return {
        "nm": (r.get("nm") or "").strip(),
        "bizNo": (r.get("bizNo") or "").strip(),
        "ceo": (r.get("ceo") or "").strip(),
        "bizType": (r.get("bizType") or "").strip(),
        "bizClass": (r.get("bizClass") or "").strip(),
        "addr": (r.get("addr") or "").strip(),
        "tel": (r.get("tel") or "").strip(),
        "fax": (r.get("fax") or "").strip(),
    }


def convert_balances(rows, clients):
    """잔액 데이터 → 거래처에 미수금/미지급금 반영"""
    cli_map = {c["nm"]: c for c in clients}
    for r in rows:
        cnm = (r.get("customerNm") or "").strip()
        balance = float(r.get("balance") or 0)
        if cnm in cli_map:
            if balance > 0:
                cli_map[cnm]["receivable"] = int(balance)
                cli_map[cnm]["cType"] = merge_customer_type(cli_map[cnm].get("cType"), "sales")
            elif balance < 0:
                cli_map[cnm]["payable"] = int(abs(balance))
                cli_map[cnm]["cType"] = merge_customer_type(cli_map[cnm].get("cType"), "purchase")


def convert_price_history(rows, clients, products):
    """price_history.csv → priceHistory KV 형식 (packflow-price-history.js 호환)"""
    if not rows:
        return []
    cli_by_nm = {c["nm"]: c for c in clients}
    prod_by_nm = {p["nm"]: p for p in products}
    history = []
    for r in rows:
        item_nm = (r.get("itemNm") or "").strip()
        cust_nm = (r.get("customerNm") or "").strip()
        if not item_nm:
            continue
        try:
            unit = float(r.get("price") or 0)
        except (ValueError, TypeError):
            unit = 0
        if not unit:
            continue
        try:
            qty = float(r.get("qty") or 0)
        except (ValueError, TypeError):
            qty = 0
        try:
            amt = float(r.get("amount") or (unit * qty))
        except (ValueError, TypeError):
            amt = unit * qty
        dt = (r.get("dt") or "").strip()
        # IoCode → refType 매핑 (1=매출, 2=매입)
        io = str(r.get("ioCode") or "").strip()
        if io == "1":
            ref_type = "sale"
        elif io == "2":
            ref_type = "purchase"
        else:
            ref_type = "almayo"
        history.append({
            "id": gid(),
            "dt": dt[:10] if dt else "",
            "tm": dt[:16] if dt else "",
            "refType": ref_type,
            "refId": "",
            "refNo": "",
            "cliNm": cust_nm,
            "prodNm": item_nm,
            "qty": qty,
            "unitPrice": int(unit),
            "amount": int(amt),
            "source": "almayo",
            "memo": (r.get("memo") or "").strip(),
            "by": "_migrate",
            "_src": "almayo"
        })
    # 최대 2000건 (packflow-price-history.js MAX 정책)
    if len(history) > 2000:
        history.sort(key=lambda h: h.get("dt", ""))
        history = history[-2000:]
    return history


def apply_recent_prices(rows, products):
    """recent_prices.csv → 각 품목의 lastPrice/lastPriceDt/lastCustomer 보강"""
    if not rows:
        return 0
    # 품목명 또는 코드로 매칭
    by_nm = {p["nm"]: p for p in products}
    by_code = {p.get("code"): p for p in products if p.get("code")}
    matched = 0
    for r in rows:
        item_nm = (r.get("itemNm") or "").strip()
        item_code = (r.get("itemCode") or "").strip()
        prod = by_nm.get(item_nm) or by_code.get(item_code)
        if not prod:
            continue
        try:
            recent_price = float(r.get("recentPrice") or 0)
        except (ValueError, TypeError):
            recent_price = 0
        if not recent_price:
            continue
        recent_dt = (r.get("recentDt") or "").strip()
        recent_cust = (r.get("recentCustomer") or "").strip()
        prod["lastPrice"] = int(recent_price)
        prod["lastPriceDt"] = recent_dt[:10] if recent_dt else ""
        prod["lastPriceCustomer"] = recent_cust
        matched += 1
    return matched


def main():
    manual_mode = "--manual" in sys.argv

    if not os.path.exists(DB_PATH):
        print(f"팩플로우 DB가 없습니다: {DB_PATH}")
        print("서버를 먼저 한 번 실행하세요: python3 server.py")
        sys.exit(1)

    db = sqlite3.connect(DB_PATH)

    # 기존 데이터 확인
    existing_cli = get_db_data(db, "cli")
    existing_prod = get_db_data(db, "prod")
    almayo_cli = [c for c in existing_cli if c.get("_src") == "almayo"]
    almayo_prod = [p for p in existing_prod if p.get("_src") == "almayo"]

    if almayo_cli or almayo_prod:
        print(f"이미 이관된 데이터가 있습니다: 거래처 {len(almayo_cli)}건, 품목 {len(almayo_prod)}건")
        ans = input("기존 이관 데이터를 삭제하고 다시 이관할까요? (y/N): ")
        if ans.lower() != "y":
            print("취소")
            db.close()
            return
        # 기존 이관 데이터 제거
        existing_cli = [c for c in existing_cli if c.get("_src") != "almayo"]
        existing_prod = [p for p in existing_prod if p.get("_src") != "almayo"]

    if manual_mode:
        print("\n=== 수동 입력 모드 ===")
        print("CSV 파일 없이 직접 입력합니다.\n")
        new_cli, new_prod = manual_input()
    else:
        # 1순위: 엑셀 파일 (.xlsx)
        xlsx_files = [f for f in os.listdir(EXPORT_DIR) if f.endswith(('.xlsx', '.xls'))] if os.path.exists(EXPORT_DIR) else []

        if xlsx_files:
            print(f"\n엑셀 파일 {len(xlsx_files)}개 발견!")
            xlsx_custs, xlsx_items = read_xlsx_auto()
            new_cli = convert_customers([{
                "nm": c["nm"], "bizNo": c.get("bizNo",""), "ceo": c.get("ceo",""),
                "bizType": c.get("bizType",""), "bizClass": c.get("bizClass",""),
                "addr": c.get("addr",""), "tel": c.get("tel",""), "fax": c.get("fax",""),
                "email": c.get("email",""), "customerType": c.get("customerType",""),
                "note": c.get("note","")
            } for c in xlsx_custs])
            new_prod = convert_items([{
                "nm": it["nm"], "code": it.get("code",""),
                "spec": it.get("spec",""), "price": str(it.get("price",0)),
                "note": it.get("note","")
            } for it in xlsx_items], new_cli)
        else:
            # 2순위: CSV
            print("\nCSV 파일 확인 중...")
            cust_rows = read_csv("customers.csv")
            item_rows = read_csv("items.csv")
            comp_rows = read_csv("company.csv")
            bal_rows = read_csv("balances.csv")
            price_history_rows = read_csv("price_history.csv")
            recent_price_rows = read_csv("recent_prices.csv")

            if not cust_rows and not item_rows:
                print("\n파일이 없습니다!")
                print(f"  폴더: {EXPORT_DIR}")
                print("\n얼마에요에서 거래처/품목 목록을 엑셀로 내보내서")
                print(f"  {EXPORT_DIR}/ 폴더에 넣어주세요.")
                print("\n또는: python3 02_import_to_packflow.py --manual")
                db.close()
                return

            new_cli = convert_customers(cust_rows)
            new_prod = convert_items(item_rows, new_cli)

            co = convert_company(comp_rows)
            if co:
                db.execute(
                    "INSERT OR REPLACE INTO data_store(key, value) VALUES('co', ?)",
                    [json.dumps(co, ensure_ascii=False)]
                )
                print(f"  회사정보: {co['nm']}")

            if bal_rows:
                convert_balances(bal_rows, new_cli)
                has_balance = sum(1 for c in new_cli if c.get("receivable") or c.get("payable"))
                print(f"  잔액 반영: {has_balance}건")

            # 최근 적용 단가 (recent_prices.csv) → prod에 lastPrice 보강
            if recent_price_rows:
                matched = apply_recent_prices(recent_price_rows, new_prod)
                print(f"  최근 단가 보강: {matched}건 / 입력 {len(recent_price_rows)}건")

            # 단가 거래 이력 (price_history.csv) → priceHistory KV 저장
            new_price_history = []
            if price_history_rows:
                new_price_history = convert_price_history(price_history_rows, new_cli, new_prod)
                print(f"  단가 이력 변환: {len(new_price_history)}건 / 원본 {len(price_history_rows)}건")

    # 기존 + 신규 합치기 (중복 이름 제거)
    existing_names = {c["nm"] for c in existing_cli}
    for c in new_cli:
        if c["nm"] not in existing_names:
            existing_cli.append(c)
            existing_names.add(c["nm"])

    existing_pnames = {p["nm"] for p in existing_prod}
    for p in new_prod:
        if p["nm"] not in existing_pnames:
            existing_prod.append(p)
            existing_pnames.add(p["nm"])

    # 저장
    set_db_data(db, "cli", existing_cli)
    set_db_data(db, "prod", existing_prod)

    # priceHistory 합쳐서 저장 (이미 있으면 _src='almayo' 만 갱신)
    added_history = 0
    if not manual_mode and 'new_price_history' in dir() and new_price_history:
        existing_ph = get_db_data(db, "priceHistory")
        # _src='almayo' 인 기존 이관 이력은 제거 (재이관 시 중복 방지)
        existing_ph = [h for h in existing_ph if h.get("_src") != "almayo"]
        merged_ph = existing_ph + new_price_history
        # 최대 2000건 (오래된 순 drop)
        if len(merged_ph) > 2000:
            merged_ph.sort(key=lambda h: h.get("dt", ""))
            merged_ph = merged_ph[-2000:]
        set_db_data(db, "priceHistory", merged_ph)
        added_history = len(new_price_history)

    db.commit()
    db.close()

    added_cli = len([c for c in existing_cli if c.get("_src") == "almayo"])
    added_prod = len([p for p in existing_prod if p.get("_src") == "almayo"])
    last_priced = len([p for p in existing_prod if p.get("lastPrice")])

    print(f"\n{'='*50}")
    print(f"  이관 완료!")
    print(f"  거래처: {added_cli}건")
    print(f"  품목  : {added_prod}건 (최근 단가: {last_priced}건)")
    if added_history:
        print(f"  단가이력: {added_history}건")
    print(f"{'='*50}")
    print(f"\n팩플로우(http://localhost:8080)에서 확인하세요.")
    print("거래처 → 거래처 목록, 품목 → 품목 목록")
    print("모바일 → 품목·단가 탭에서 '최근 적용 단가' 확인")


def manual_input():
    """대화식 수동 입력"""
    clients = []
    products = []

    print("--- 거래처 입력 (빈 줄 입력 시 종료) ---")
    while True:
        nm = input("거래처명: ").strip()
        if not nm:
            break
        tel = input("  전화: ").strip()
        addr = input("  주소: ").strip()
        tp = input("  유형 (sales/purchase/both) [sales]: ").strip() or "sales"
        clients.append({
            "id": gid(), "nm": nm, "tel": tel, "addr": addr,
            "cType": tp, "bizNo": "", "ceo": "", "fax": "", "email": "",
            "ps": "", "note": "", "cat": now_str(), "_src": "almayo"
        })
        print(f"  → {nm} 추가됨")

    print(f"\n거래처 {len(clients)}건 입력 완료\n")

    print("--- 품목 입력 (빈 줄 입력 시 종료) ---")
    default_procs = [
        {"nm": "인쇄", "tp": "out"},
        {"nm": "코팅", "tp": "in"},
        {"nm": "톰슨", "tp": "in"},
        {"nm": "접착", "tp": "in"},
    ]
    while True:
        nm = input("품목명: ").strip()
        if not nm:
            break
        spec = input("  규격: ").strip()
        price = input("  단가: ").strip()
        products.append({
            "id": gid(), "nm": nm, "code": "", "cnm": "", "cid": "",
            "spec": spec, "price": int(price) if price else 0,
            "paper": "", "papers": [], "fabric": "", "fabrics": [],
            "procs": default_procs, "ps": "", "gold": "", "mold": "",
            "hand": "", "nt": "", "caut": "", "cat": now_str(), "_src": "almayo"
        })
        print(f"  → {nm} 추가됨")

    print(f"\n품목 {len(products)}건 입력 완료")
    return clients, products


if __name__ == "__main__":
    main()
